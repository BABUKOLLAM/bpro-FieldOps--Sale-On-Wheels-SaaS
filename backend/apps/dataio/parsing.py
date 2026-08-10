import csv
import io

from openpyxl import load_workbook


def parse_upload(django_file) -> list[dict]:
    name = (django_file.name or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(django_file)
    return _parse_csv(django_file)


def _parse_csv(django_file) -> list[dict]:
    text = django_file.read().decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def _parse_xlsx(django_file) -> list[dict]:
    wb = load_workbook(django_file, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        rows.append({headers[i]: ("" if row[i] is None else row[i]) for i in range(len(headers))})
    return rows
