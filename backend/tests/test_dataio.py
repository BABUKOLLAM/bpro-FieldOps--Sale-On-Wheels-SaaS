"""apps.dataio: bulk import/export for master data (products, price list
entries, staff, routes, customers, vehicles, godowns). Covers the generic
upsert engine, the two entities with bespoke row handling (staff, prices),
permission gating per entity, and that a password hash never leaves the
system via export."""

import csv
import io

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import Role, User, UserRole
from apps.catalog.models import Item, ItemCategory, PriceList, PriceListItem, UOM
from apps.customers.models import Beat, Customer


def _csv_upload(headers, rows, filename="upload.csv"):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    from django.core.files.uploadedfile import SimpleUploadedFile

    return SimpleUploadedFile(filename, buf.getvalue().encode("utf-8"), content_type="text/csv")


@pytest.fixture
def uom(db):
    return UOM.objects.create(code="PCS", name="Pieces")


@pytest.fixture
def category(db):
    return ItemCategory.objects.create(name="Beverages")


@pytest.fixture
def product(uom, category):
    return Item.objects.create(sku="SKU-1", name="Cola", base_uom=uom, category=category)


# ---- export ----


@pytest.mark.django_db
def test_export_xlsx_contains_product_rows(admin, product):
    client = APIClient()
    client.force_authenticate(user=admin)
    response = client.get("/api/dataio/entities/products/export/?filetype=xlsx")
    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][:2] == ("sku", "name")
    assert any(r[0] == "SKU-1" for r in rows[1:])


@pytest.mark.django_db
def test_export_csv_format(admin, product):
    client = APIClient()
    client.force_authenticate(user=admin)
    response = client.get("/api/dataio/entities/products/export/?filetype=csv")
    assert response.status_code == 200
    assert response["Content-Type"] == "text/csv"
    text = response.content.decode("utf-8-sig")
    reader = list(csv.reader(io.StringIO(text)))
    assert reader[0][:2] == ["sku", "name"]


@pytest.mark.django_db
def test_staff_export_never_includes_password_hash(admin, agent):
    client = APIClient()
    client.force_authenticate(user=admin)
    response = client.get("/api/dataio/entities/staff/export/?filetype=csv")
    assert response.status_code == 200
    text = response.content.decode("utf-8-sig")
    header = text.splitlines()[0]
    assert "password" not in header.split(",")
    assert "pbkdf2" not in text  # Django's password hash prefix must never appear


@pytest.mark.django_db
def test_template_download_has_all_columns_including_password(admin):
    client = APIClient()
    client.force_authenticate(user=admin)
    response = client.get("/api/dataio/entities/staff/template/?filetype=csv")
    header = response.content.decode("utf-8-sig").splitlines()[0]
    assert "password" in header.split(",")


# ---- permission gating ----


@pytest.mark.django_db
def test_export_denied_without_permission(agent):
    client = APIClient()
    client.force_authenticate(user=agent)
    response = client.get("/api/dataio/entities/products/export/")
    assert response.status_code == 403


@pytest.mark.django_db
def test_unknown_entity_404s(admin):
    client = APIClient()
    client.force_authenticate(user=admin)
    response = client.get("/api/dataio/entities/not-a-real-entity/export/")
    assert response.status_code == 404


# ---- generic import (products) ----


@pytest.mark.django_db
def test_import_creates_and_updates_with_fk_resolution(admin, uom, category, product):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["sku", "name", "category", "base_uom", "hsn_code", "gst_rate", "barcode", "is_active"],
        [
            ["SKU-1", "Cola Renamed", "Beverages", "PCS", "2202", "18", "", "true"],
            ["SKU-2", "Fanta", "Beverages", "PCS", "2202", "18", "", "true"],
        ],
    )
    response = client.post("/api/dataio/entities/products/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200, response.data
    assert response.data["created_count"] == 1
    assert response.data["updated_count"] == 1
    assert response.data["errors"] == []
    product.refresh_from_db()
    assert product.name == "Cola Renamed"
    assert Item.objects.filter(sku="SKU-2").exists()


@pytest.mark.django_db
def test_import_reports_row_errors_without_aborting_whole_file(admin, uom, category):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["sku", "name", "category", "base_uom", "hsn_code", "gst_rate", "barcode", "is_active"],
        [
            ["", "Missing SKU", "Beverages", "PCS", "", "", "", "true"],
            ["SKU-BAD", "Bad UOM", "Beverages", "NOPE", "", "", "", "true"],
            ["SKU-OK", "Good Row", "Beverages", "PCS", "", "", "", "true"],
        ],
    )
    response = client.post("/api/dataio/entities/products/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["created_count"] == 1
    assert len(response.data["errors"]) == 2
    assert response.data["errors"][0]["row"] == 2
    assert Item.objects.filter(sku="SKU-OK").exists()
    assert not Item.objects.filter(sku="SKU-BAD").exists()


@pytest.mark.django_db
def test_import_auto_creates_missing_category(admin, uom):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["sku", "name", "category", "base_uom", "hsn_code", "gst_rate", "barcode", "is_active"],
        [["SKU-NEW", "New Item", "Brand New Category", "PCS", "", "", "", "true"]],
    )
    response = client.post("/api/dataio/entities/products/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["created_count"] == 1
    assert ItemCategory.objects.filter(name="Brand New Category").exists()


@pytest.mark.django_db
def test_import_blank_optional_text_field_does_not_crash(admin, uom, category):
    """Regression: Item.barcode/hsn_code are CharField(blank=True) with no
    null=True — coercing a blank cell to None used to violate the NOT NULL
    constraint at the DB layer."""
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["sku", "name", "category", "base_uom", "hsn_code", "gst_rate", "barcode", "is_active"],
        [["SKU-BLANK", "Blank Fields", "Beverages", "PCS", "", "", "", "true"]],
    )
    response = client.post("/api/dataio/entities/products/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["errors"] == []
    saved = Item.objects.get(sku="SKU-BLANK")
    assert saved.barcode == ""
    assert saved.hsn_code == ""


@pytest.mark.django_db
def test_import_creates_dataio_job_audit_record(admin, uom, category):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["sku", "name", "category", "base_uom", "hsn_code", "gst_rate", "barcode", "is_active"],
        [["SKU-JOB", "Job Test", "Beverages", "PCS", "", "", "", "true"]],
    )
    client.post("/api/dataio/entities/products/import/", {"file": upload}, format="multipart")
    jobs = client.get("/api/dataio/jobs/?entity_slug=products")
    assert jobs.status_code == 200
    assert any(j["created_count"] == 1 for j in jobs.data["results"])


@pytest.mark.django_db
def test_jobs_list_hides_entities_the_user_cannot_manage(admin, agent, uom, category):
    admin_client = APIClient()
    admin_client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["sku", "name", "category", "base_uom", "hsn_code", "gst_rate", "barcode", "is_active"],
        [["SKU-SCOPED", "Scoped Test", "Beverages", "PCS", "", "", "", "true"]],
    )
    admin_client.post("/api/dataio/entities/products/import/", {"file": upload}, format="multipart")

    agent_client = APIClient()
    agent_client.force_authenticate(user=agent)
    jobs = agent_client.get("/api/dataio/jobs/")
    assert jobs.status_code == 200
    assert all(j["entity_slug"] != "products" for j in jobs.data["results"])


# ---- staff (bespoke row handling) ----


@pytest.mark.django_db
def test_staff_import_sets_password_only_on_create(admin):
    Role.seed_defaults()
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["username", "first_name", "last_name", "email", "phone", "employee_code",
         "is_field_agent", "reporting_manager", "is_active", "role", "password"],
        [["newhire@test.local", "New", "Hire", "", "", "", "true", "", "true", "van_salesman", "TempPass1"]],
    )
    response = client.post("/api/dataio/entities/staff/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["created_count"] == 1
    user = User.objects.get(username="newhire@test.local")
    assert user.check_password("TempPass1")
    assert UserRole.objects.filter(user=user, role__name="van_salesman").exists()


@pytest.mark.django_db
def test_staff_import_blank_password_does_not_clear_existing(admin, agent):
    before_hash = agent.password
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["username", "first_name", "last_name", "email", "phone", "employee_code",
         "is_field_agent", "reporting_manager", "is_active", "role", "password"],
        [[agent.username, "Updated", "Name", "", "", "", "true", "", "true", "", ""]],
    )
    response = client.post("/api/dataio/entities/staff/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["updated_count"] == 1
    agent.refresh_from_db()
    assert agent.password == before_hash
    assert agent.first_name == "Updated"


@pytest.mark.django_db
def test_staff_import_unknown_role_is_reported_as_row_error(admin):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["username", "first_name", "last_name", "email", "phone", "employee_code",
         "is_field_agent", "reporting_manager", "is_active", "role", "password"],
        [["bad-role@test.local", "", "", "", "", "", "false", "", "true", "not_a_real_role", ""]],
    )
    response = client.post("/api/dataio/entities/staff/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert len(response.data["errors"]) == 1
    assert "not_a_real_role" in response.data["errors"][0]["message"]


# ---- price list entries (bespoke row handling) ----


@pytest.mark.django_db
def test_price_import_auto_creates_price_list_and_links_by_sku(admin, product):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(["price_list", "sku", "rate"], [["Wholesale", "SKU-1", "15.50"]])
    response = client.post("/api/dataio/entities/prices/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["created_count"] == 1
    assert PriceList.objects.filter(name="Wholesale").exists()
    entry = PriceListItem.objects.get(price_list__name="Wholesale", item__sku="SKU-1")
    assert str(entry.rate) == "15.50"


@pytest.mark.django_db
def test_price_import_unknown_sku_is_row_error(admin):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(["price_list", "sku", "rate"], [["Wholesale", "NOPE", "10"]])
    response = client.post("/api/dataio/entities/prices/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["created_count"] == 0
    assert "NOPE" in response.data["errors"][0]["message"]


@pytest.mark.django_db
def test_price_import_upserts_on_second_run(admin, product):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload1 = _csv_upload(["price_list", "sku", "rate"], [["Wholesale", "SKU-1", "10"]])
    client.post("/api/dataio/entities/prices/import/", {"file": upload1}, format="multipart")
    upload2 = _csv_upload(["price_list", "sku", "rate"], [["Wholesale", "SKU-1", "12.25"]])
    response = client.post("/api/dataio/entities/prices/import/", {"file": upload2}, format="multipart")
    assert response.data["updated_count"] == 1
    assert response.data["created_count"] == 0
    entry = PriceListItem.objects.get(price_list__name="Wholesale", item__sku="SKU-1")
    assert str(entry.rate) == "12.25"


# ---- routes/customers upsert-by-practical-key ----


@pytest.mark.django_db
def test_routes_import_upserts_by_name(admin, agent):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(["name", "assigned_agent", "is_active"], [["North Zone", agent.username, "true"]])
    response = client.post("/api/dataio/entities/routes/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["created_count"] == 1
    beat = Beat.objects.get(name="North Zone")
    assert beat.assigned_agent == agent


@pytest.mark.django_db
def test_customers_import_creates_with_defaults(admin):
    client = APIClient()
    client.force_authenticate(user=admin)
    upload = _csv_upload(
        ["code", "name", "category", "gstin", "phone", "credit_limit", "credit_days", "is_blocked", "is_active"],
        [["CUST-IMP", "Imported Store", "", "", "", "", "", "", "true"]],
    )
    response = client.post("/api/dataio/entities/customers/import/", {"file": upload}, format="multipart")
    assert response.status_code == 200
    assert response.data["created_count"] == 1
    customer = Customer.objects.get(code="CUST-IMP")
    assert customer.credit_limit == 0
    assert customer.gstin == ""
